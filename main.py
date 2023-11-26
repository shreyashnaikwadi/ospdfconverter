import os
import io
import shutil
import time
from tkinter import Tk, filedialog, simpledialog, messagebox
from tkinter import *
from PIL import Image, ImageTk
import customtkinter
from CTkMessagebox import CTkMessagebox
import openpyxl
import pandas as pd
from fpdf import FPDF
import comtypes.client
import PyPDF2  # Add this import
from docx2pdf import convert as convert_to_pdf
from img2pdf import convert as image_to_pdf
from pptx import Presentation
from reportlab.pdfgen import canvas
from PyPDF2 import PdfReader, PdfWriter
from pdf2image import convert_from_path, convert_from_bytes
from pdf2docx import Converter


def main():


    def open_file(filetypes):
        file_path = filedialog.askopenfilename(filetypes=filetypes)
        return file_path

    def save_file(filetypes):
        file_path = filedialog.asksaveasfilename(defaultextension=filetypes[0][1], filetypes=filetypes)
        return file_path


    # CONVERT PDF TO WORD
    def convert_pdf_to_word():
        input_pdf = filedialog.askopenfilename(filetypes=[("PDF Files", "*.pdf")])
        if input_pdf:
            output_word = filedialog.asksaveasfilename(defaultextension=".docx", filetypes=[("Word Files", "*.docx")])
            if output_word:
                try:
                    # Convert PDF to Word using pdf2docx library
                    cv = Converter(input_pdf)
                    cv.convert(output_word, start=0, end=None)
                    cv.close()
                    CTkMessagebox(title="Conversion Complete", message="PDF to Word conversion is complete!", icon="check", option_1="Ok")
                except Exception as e:
                    messagebox.showerror("Error", f"An error occurred: {e}")
                    CTkMessagebox(title="Error", message="Something went wrong!!!", icon="cancel")

    # CONVERT WORD TO PDF 
    def convert_word_to_pdf():
        input_file = open_file([("Word Files", "*.docx")])
        if input_file:
            output_file = save_file([("PDF Files", "*.pdf")])
            if output_file:
                # Code to convert Word to PDF using external library here
                convert_to_pdf(input_file, output_file)
                CTkMessagebox(title="Conversion Complete", message="Word to PDF conversion is complete!", icon="check", option_1="Ok")
            else:
                CTkMessagebox(title="Error", message="Something went wrong!!!", icon="cancel")


    # CONVERT IMAGE TO PDF
    def convert_image_to_pdf():
        input_file = open_file([("Image Files", "*.png;*.jpg;*.jpeg")])
        if input_file:
            output_file = save_file([("PDF Files", "*.pdf")])
            if output_file:
                # Code to convert Image to PDF using external library here
                with open(output_file, "wb") as f:
                    f.write(image_to_pdf([input_file]))
                CTkMessagebox(title="Conversion Complete", message="Image to PDF conversion is complete!", icon="check", option_1="Ok")
            else:
                CTkMessagebox(title="Error", message="Something went wrong!!!", icon="cancel")


    # CONVERT PDF TO IMAGE
    def pdf_to_image():
        pdf_file = filedialog.askopenfilename(filetypes=[("PDF Files", "*.pdf")])
        if pdf_file:
            try:
                output_directory = filedialog.askdirectory(title="Select Output Directory")
                if output_directory:
                    convert_pdf_to_images(pdf_file, output_directory)
                    CTkMessagebox(title="Conversion Complete", message="PDF to Image conversion successful!", icon="check", option_1="Ok")
            except Exception as e:
                CTkMessagebox(title="Error", message="Something went wrong!!!", icon="cancel")


    def convert_pdf_to_images(pdf_file, output_directory):
        images = convert_from_path(pdf_file)

        for i, image in enumerate(images):
            image_path = os.path.join(output_directory, f"page_{i + 1}.png")
            image.save(image_path, "PNG")



    # CONVERT  EXCEL TO PDF
    def convert_excel_to_pdf():
        excel_file = open_file([("Excel Files", "*.xlsx")])
        if excel_file:
            try:
                output_path = save_file([("PDF Files", "*.pdf")])
                if output_path:
                    #  a new PDF using fpdf
                    pdf = FPDF()
                    pdf.add_page()

                    # Set font (e.g., Arial) and font size
                    pdf.set_font("Arial", size=12)

                    # Open the Excel file and read its content
                    xls = openpyxl.load_workbook(excel_file)
                    sheet = xls.active
                    max_rows = sheet.max_row
                    max_cols = sheet.max_column

                    # Loop through the cells and add them to the PDF
                    for row in range(1, max_rows + 1):
                        for col in range(1, max_cols + 1):
                            cell_value = sheet.cell(row=row, column=col).value
                            pdf.cell(200, 10, str(cell_value), ln=True)

                    # Save the PDF
                    pdf.output(output_path)
                    CTkMessagebox(title="Conversion Complete", message="Excel successfully converted to PDF!", icon="check", option_1="Ok")
            except Exception as e:
                messagebox.showerror("Error", f"An error occurred: {e}")
                CTkMessagebox(title="Error", message="Something went wrong!!!", icon="cancel")

    # CONVERT PDF TO EXCEL 
    def convert_pdf_to_excel(pdf_path, excel_path):
        pdf_text = ""
        with open(pdf_path, 'rb') as pdf_file:
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            for page in pdf_reader.pages:
                pdf_text += page.extract_text()

        data = [line.split('\t') for line in pdf_text.split('\n')]

        df = pd.DataFrame(data)
        df.to_excel(excel_path, index=False, header=False)
        CTkMessagebox(title="Conversion Complete", message="PDF successfully converted to Excel!", icon="check", option_1="Ok")

    # Function to handle the PDF to Excel button
    def convert_pdf_to_excel_callback():
        pdf_file = open_file([("PDF Files", "*.pdf")])
        if pdf_file:
            try:
                output_path = save_file([("Excel Files", "*.xlsx")])
                if output_path:
                    convert_pdf_to_excel(pdf_file, output_path)
            except Exception as e:
                CTkMessagebox(title="Error", message="Something went wrong!!!", icon="cancel")


    # PROTECT PDF 
    def protect_pdf(input_pdf, output_pdf, password):
        pdf_reader = PdfReader(input_pdf)
        pdf_writer = PdfWriter()

        for page in pdf_reader.pages:
            pdf_writer.add_page(page)

        pdf_writer.encrypt(user_pwd=password, owner_pwd=None, use_128bit=True)

        with open(output_pdf, "wb") as output_file:
            pdf_writer.write(output_file)
    def protect_pdf_callback():
        pdf_file = open_file([("PDF Files", "*.pdf")])
        if pdf_file:
            try:
                password = simpledialog.askstring("Password Protection", "Enter password:")
                if password:
                    output_path = save_file([("PDF Files", "*.pdf")])
                    if output_path:
                        protect_pdf(pdf_file, output_path, password)
                        CTkMessagebox(title="PDF Protected", message="PDF file protected with password successfully!", icon="check", option_1="Ok")
            except Exception as e:
                CTkMessagebox(title="Error", message="Something went wrong!!!", icon="cancel")


    # ORGANIZE PDF BY DATE
    def organize_pdfs(input_directory, output_directory):
        pdf_files = [f for f in os.listdir(input_directory) if f.endswith(".pdf")]

        for pdf_file in pdf_files:
            pdf_path = os.path.join(input_directory, pdf_file)
            creation_time = os.path.getctime(pdf_path)

            year = str(time.localtime(creation_time).tm_year)
            month = str(time.localtime(creation_time).tm_mon).zfill(2)  # Zero padding
            destination_dir = os.path.join(output_directory, year, month)

            os.makedirs(destination_dir, exist_ok=True)
            shutil.move(pdf_path, os.path.join(destination_dir, pdf_file))

        CTkMessagebox(title="PDF Organized", message="PDF files organized by date successfully!", icon="check", option_1="Ok")
    def organize_pdfs_callback():
        input_directory = filedialog.askdirectory(title="Select Input Directory")
        if input_directory:
            output_directory = filedialog.askdirectory(title="Select Output Directory")
            if output_directory:
                organize_pdfs(input_directory, output_directory)

    customtkinter.set_appearance_mode("System")
    customtkinter.set_default_color_theme("blue")

    root = customtkinter.CTk()
    root.title("OS PDF CONVERTER")
    root.geometry("800x600")
    root.resizable(width=False, height=False)

    # Label OS PDF CONVERTER

    title_label = customtkinter.CTkLabel(root, text="OS PDF CONVERTER", font=("Ariel bold", 36))
    title_label.place(relx=0.5, rely=0.1, anchor=N)

    # Frame

    button_frame = customtkinter.CTkFrame(master=root, width=400, height=300)
    button_frame.place(relx=0.5, rely=0.5, anchor=CENTER)


    convert_pdf_to_word_button = customtkinter.CTkButton(master=button_frame, text="PDF to Word", command=convert_pdf_to_word)
    convert_pdf_to_word_button.grid(row=0, column=0, padx=10, pady=10)

    convert_word_to_pdf_button = customtkinter.CTkButton(master=button_frame, text="Word to PDF", command=convert_word_to_pdf)
    convert_word_to_pdf_button.grid(row=0, column=1, padx=10, pady=10)

    convert_image_to_pdf_button = customtkinter.CTkButton(master=button_frame, text="Image to PDF", command=convert_image_to_pdf)
    convert_image_to_pdf_button.grid(row=1, column=0, padx=10, pady=10)

    convert_excel_to_pdf_button = customtkinter.CTkButton(master=button_frame, text="Excel to PDF", command=convert_excel_to_pdf)
    convert_excel_to_pdf_button.grid(row=1, column=1, padx=10, pady=10)

    convert_pdf_to_image_button = customtkinter.CTkButton(master=button_frame, text="PDF to Image", command=pdf_to_image)
    convert_pdf_to_image_button.grid(row=2, column=0, padx=10, pady=10)

    convert_pdf_to_excel_button = customtkinter.CTkButton(master=button_frame, text="PDF to Excel", command=convert_pdf_to_excel_callback)
    convert_pdf_to_excel_button.grid(row=2, column=1, padx=10, pady=10)

    protect_pdf_button = customtkinter.CTkButton(master=button_frame, text="Protect PDF", command=protect_pdf_callback)
    protect_pdf_button.grid(row=3, column=0, padx=10, pady=10)

    organize_pdfs_button = customtkinter.CTkButton(master=button_frame, text="Organize PDFs", command=organize_pdfs_callback)
    organize_pdfs_button.grid(row=3, column=1, padx=10, pady=10)

    root.mainloop()

if __name__ == "__main__":
    main()